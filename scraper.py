import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

URL = "https://www.worldometers.info/coronavirus/"
page = requests.get(URL)

soup = BeautifulSoup(page.content, "html.parser")
table = soup.find("table", id="main_table_countries_today")
all_bodies = soup.findAll("tbody")
main_body = all_bodies[0]
main_rows = main_body.findAll("tr")

continents = []
total_numbers = []
for index_x in range(0, 6):
    continents.append(main_rows[index_x].findAll("td")[1].text.strip())
    total_numbers.append(main_rows[index_x].findAll("td")[2].text.strip())

wb = Workbook()
ws = wb.active
ws.title = "Corona cases"

for x_index, _ in enumerate(continents):
    ws['A' + str(x_index + 1)] = continents[x_index]
    ws['B' + str(x_index + 1)] = total_numbers[x_index]

wb.save('cases.xlsx')
